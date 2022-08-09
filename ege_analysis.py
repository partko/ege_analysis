import openpyxl
from openpyxl.styles import numbers
from openpyxl.styles.numbers import BUILTIN_FORMATS
from openpyxl.styles import Alignment


while True:
    try:
        wb = openpyxl.load_workbook("input.xlsx") #wb - workbook
        break
    except:
        print("В папке с программой должен лежать файл с входными данными \"input.xlsx\". Такой файл не найден.")
        a = input()


input_sheet_name = 'результаты оценивания _Х'

while True:
    try:
        ws = wb[input_sheet_name] #ws - worksheet
        break
    except:
        print("В файле должен быть лист с именем \"результаты оценивания _Х\". Если лист называется по-другому, скопируйте его название сюда (чтобы вставить - правая кнопка мыши): ")
        input_sheet_name = input()


column_C = ws['C']
column_I = ws['I']

column_J = ws['J']
column_H = ws['H']

column_N = ws['N']
column_O = ws['O']
column_P = ws['P']
column_Q = ws['Q'] 
column_R = ws['R'] 
column_S = ws['S']

C_tasks = [column_N, column_O, column_P, column_Q, column_R, column_S]

FIO = []

for i in range(6, len(column_J)):
    #print(column_J[i].value)
    if column_J[i].value not in FIO and column_J[i].value != None:
        FIO.append(column_J[i].value)

#print(FIO)

total_amount = [] #Общая сумма проверенных работ на каждого преподавателя

for i in range(len(FIO)):
    total_amount.append(0)

for j, elements in enumerate(FIO):
    for i in range(6, len(column_J)):
        if elements == column_J[i].value:
            total_amount[j] +=1

#print(total_amount)


#total_amount_without_3 = []
not_included_in_the_3rd_check = []
not_included_in_the_3rd_check_with_discrepancies_1 = []
not_included_in_the_3rd_check_with_discrepancies_2 = []
not_included_in_the_3rd_check_with_discrepancies_3_and_more = []
not_included_in_the_3rd_check_with_discrepancies_overstatement = []
not_included_in_the_3rd_check_with_discrepancies_understatement = []


included_in_the_3rd_check = []

included_in_the_3rd_check_with_a_discrepancy_2 = []
included_in_the_3rd_check_with_a_discrepancy_3 = []
included_in_the_3rd_check_with_a_discrepancy_4 = []
included_in_the_3rd_check_with_a_discrepancy_5_and_more = []
included_in_the_3rd_check_with_a_discrepancy_X_mark = []

discrepancies_with_the_3rd_expert = []
discrepancies_with_the_3rd_expert_overstatement = []
discrepancies_with_the_3rd_expert_understatement = []
unappreciated_tasks = []


for i in range(len(FIO)):
    #total_amount_without_3.append(0)
    not_included_in_the_3rd_check.append(0)
    not_included_in_the_3rd_check_with_discrepancies_1.append(0)
    not_included_in_the_3rd_check_with_discrepancies_2.append(0)
    not_included_in_the_3rd_check_with_discrepancies_3_and_more.append(0)
    not_included_in_the_3rd_check_with_discrepancies_overstatement.append(0)
    not_included_in_the_3rd_check_with_discrepancies_understatement.append(0)


    included_in_the_3rd_check.append(0)
    included_in_the_3rd_check_with_a_discrepancy_2.append(0)
    included_in_the_3rd_check_with_a_discrepancy_3.append(0)
    included_in_the_3rd_check_with_a_discrepancy_4.append(0)
    included_in_the_3rd_check_with_a_discrepancy_5_and_more.append(0)
    included_in_the_3rd_check_with_a_discrepancy_X_mark.append(0)

    discrepancies_with_the_3rd_expert.append(0)
    discrepancies_with_the_3rd_expert_overstatement.append(0)
    discrepancies_with_the_3rd_expert_understatement.append(0)
    unappreciated_tasks.append(0)


is_the_top = False
count_cells = 1
discrepancies = 0
is_not_3_overstatement = False
is_not_3_understatement = False

counter = 0
is_the_X_mark = False

is_the_discrepancies_with_the_3rd_expert = False
is_the_discrepancies_with_the_3rd_expert_overstatement = False
is_the_discrepancies_with_the_3rd_expert_understatement = False
is_the_unappreciated_tasks = False

for j, elements in enumerate(FIO):
    for i in range(6, len(column_J)):
        if elements == column_J[i].value and column_H[i].value == None: #Работы, НЕ попавшие на 3-ю проверку
            not_included_in_the_3rd_check[j] +=1
            #total_amount_without_3[j] +=1
            if column_I[i].value == 0:
                pass
            else:
                if column_C[i].value == column_C[i+1].value:
                    if column_C[i].value != column_C[i-1].value:
                        for elem in C_tasks:
                            if elem[i].value == 'X' and elem[i+1].value == 'X':
                                pass
                            elif elem[i].value == 'X' and elem[i+1].value == 0:
                                pass
                            elif elem[i].value == 0 and elem[i+1].value == 'X':
                                pass
                            elif elem[i].value == 'X' and elem[i+1].value != 'X' and elem[i+1].value != 0:
                                discrepancies += elem[i+1].value
                                is_not_3_understatement = True
                            elif elem[i].value != 'X' and elem[i].value != 0 and elem[i+1].value == 'X':
                                discrepancies += elem[i].value
                                is_not_3_overstatement = True
                            elif elem[i].value > elem[i+1].value:
                                discrepancies += abs(elem[i].value - elem[i+1].value)
                                is_not_3_overstatement = True
                            elif elem[i].value < elem[i+1].value:
                                discrepancies += abs(elem[i].value - elem[i+1].value)
                                is_not_3_understatement = True



                    else:
                        while not is_the_top:
                            if column_C[i].value == column_C[i-count_cells].value:
                                count_cells += 1
                            else:
                                is_the_top = True
                        if count_cells % 2 == 0:
                            for elem in C_tasks:
                                if elem[i].value == 'X' and elem[i-1].value == 'X':
                                    pass
                                elif elem[i].value == 'X' and elem[i-1].value == 0:
                                    pass
                                elif elem[i].value == 0 and elem[i-1].value == 'X':
                                    pass
                                elif elem[i].value == 'X' and elem[i-1].value != 'X' and elem[i-1].value != 0:
                                    discrepancies += elem[i-1].value
                                    is_not_3_understatement = True
                                elif elem[i].value != 'X' and elem[i].value != 0 and elem[i-1].value == 'X':
                                    discrepancies += elem[i].value
                                    is_not_3_overstatement = True
                                elif elem[i].value > elem[i-1].value:
                                    discrepancies += abs(elem[i].value - elem[i-1].value)
                                    is_not_3_overstatement = True
                                elif elem[i].value < elem[i-1].value:
                                    discrepancies += abs(elem[i].value - elem[i-1].value)
                                    is_not_3_understatement = True
                        else:
                            for elem in C_tasks:
                                if elem[i].value == 'X' and elem[i+1].value == 'X':
                                    pass
                                elif elem[i].value == 'X' and elem[i+1].value == 0:
                                    pass
                                elif elem[i].value == 0 and elem[i+1].value == 'X':
                                    pass
                                elif elem[i].value == 'X' and elem[i+1].value != 'X' and elem[i+1].value != 0:
                                    discrepancies += elem[i+1].value
                                    is_not_3_understatement = True
                                elif elem[i].value != 'X' and elem[i].value != 0 and elem[i+1].value == 'X':
                                    discrepancies += elem[i].value
                                    is_not_3_overstatement = True
                                elif elem[i].value > elem[i+1].value:
                                    discrepancies += abs(elem[i].value - elem[i+1].value)
                                    is_not_3_overstatement = True
                                elif elem[i].value < elem[i+1].value:
                                    discrepancies += abs(elem[i].value - elem[i+1].value)
                                    is_not_3_understatement = True



                elif column_C[i].value == column_C[i-1].value:
                    if column_C[i].value != column_C[i+1].value:
                        for elem in C_tasks:
                            if elem[i].value == 'X' and elem[i-1].value == 'X':
                                pass
                            elif elem[i].value == 'X' and elem[i-1].value == 0:
                                pass
                            elif elem[i].value == 0 and elem[i-1].value == 'X':
                                pass
                            elif elem[i].value == 'X' and elem[i-1].value != 'X' and elem[i-1].value != 0:
                                discrepancies += elem[i-1].value
                                is_not_3_understatement = True
                            elif elem[i].value != 'X' and elem[i].value != 0 and elem[i-1].value == 'X':
                                discrepancies += elem[i].value
                                is_not_3_overstatement = True
                            elif elem[i].value > elem[i-1].value:
                                discrepancies += abs(elem[i].value - elem[i-1].value)
                                is_not_3_overstatement = True
                            elif elem[i].value < elem[i-1].value:
                                discrepancies += abs(elem[i].value - elem[i-1].value)
                                is_not_3_understatement = True


                    else:
                        while not is_the_top:
                            if column_C[i].value == column_C[i-count_cells].value:
                                count_cells += 1
                            else:
                                is_the_top = True
                        if count_cells % 2 == 0:
                            for elem in C_tasks:
                                if elem[i].value == 'X' and elem[i-1].value == 'X':
                                    pass
                                elif elem[i].value == 'X' and elem[i-1].value == 0:
                                    pass
                                elif elem[i].value == 0 and elem[i-1].value == 'X':
                                    pass
                                elif elem[i].value == 'X' and elem[i-1].value != 'X' and elem[i-1].value != 0:
                                    discrepancies += elem[i-1].value
                                    is_not_3_understatement = True
                                elif elem[i].value != 'X' and elem[i].value != 0 and elem[i-1].value == 'X':
                                    discrepancies += elem[i].value
                                    is_not_3_overstatement = True
                                elif elem[i].value > elem[i-1].value:
                                    discrepancies += abs(elem[i].value - elem[i-1].value)
                                    is_not_3_overstatement = True
                                elif elem[i].value < elem[i-1].value:
                                    discrepancies += abs(elem[i].value - elem[i-1].value)
                                    is_not_3_understatement = True
                        else:
                            for elem in C_tasks:
                                if elem[i].value == 'X' and elem[i+1].value == 'X':
                                    pass
                                elif elem[i].value == 'X' and elem[i+1].value == 0:
                                    pass
                                elif elem[i].value == 0 and elem[i+1].value == 'X':
                                    pass
                                elif elem[i].value == 'X' and elem[i+1].value != 'X' and elem[i+1].value != 0:
                                    discrepancies += elem[i+1].value
                                    is_not_3_understatement = True
                                elif elem[i].value != 'X' and elem[i].value != 0 and elem[i+1].value == 'X':
                                    discrepancies += elem[i].value
                                    is_not_3_overstatement = True
                                elif elem[i].value > elem[i+1].value:
                                    discrepancies += abs(elem[i].value - elem[i+1].value)
                                    is_not_3_overstatement = True
                                elif elem[i].value < elem[i+1].value:
                                    discrepancies += abs(elem[i].value - elem[i+1].value)
                                    is_not_3_understatement = True


            is_the_top = False
            count_cells = 1
            if discrepancies == 1:
                not_included_in_the_3rd_check_with_discrepancies_1[j] += 1
            elif discrepancies == 2:
                not_included_in_the_3rd_check_with_discrepancies_2[j] += 1
            elif discrepancies >= 3:
                not_included_in_the_3rd_check_with_discrepancies_3_and_more[j] += 1

            if is_not_3_overstatement:
                not_included_in_the_3rd_check_with_discrepancies_overstatement[j] += 1
                is_not_3_overstatement = False
            if is_not_3_understatement:
                not_included_in_the_3rd_check_with_discrepancies_understatement[j] += 1
                is_not_3_understatement = False

            discrepancies = 0




        #Работы, попавшие на 3-ю проверку (1 и 2 эксперт)
        elif elements == column_J[i].value and column_H[i].value != None and (column_H[i-2].value == None or column_H[i-2].value != column_H[i].value):
            included_in_the_3rd_check[j] +=1
            #total_amount_without_3[j] +=1

            for elem in C_tasks:
                if column_H[i-1].value == None or column_H[i-1].value != column_H[i].value:
                    if elem[i].value == 'X' and elem[i+1].value == 'X':
                        pass
                    elif elem[i].value == 'X' and elem[i+1].value == 0:
                        pass
                    elif elem[i].value == 0 and elem[i+1].value == 'X':
                        pass
                    elif elem[i].value == 'X' and elem[i+1].value != 'X' and elem[i+1].value != 0:
                        is_the_X_mark = True
                        #print(elements, i)
                        #counter += elem[i+1].value #X-задания не должны попадать в другие столбцы
                        
                    elif elem[i].value != 'X' and elem[i].value != 0 and elem[i+1].value == 'X':
                        counter += elem[i].value

                    elif elem[i].value != elem[i+1].value:
                        #print(elem[i].value)
                        #print(elem[i+1].value)
                        counter += abs(elem[i].value - elem[i+1].value)


                    if elem[i+2].value == 'X':
                        pass
                    elif elem[i+2].value != 'X':
                        if elem[i].value == elem[i+2].value:
                            pass
                        elif elem[i].value == 'X':
                            is_the_discrepancies_with_the_3rd_expert = True
                            is_the_unappreciated_tasks = True
                        elif elem[i].value > elem[i+2].value:
                            is_the_discrepancies_with_the_3rd_expert = True
                            is_the_discrepancies_with_the_3rd_expert_overstatement = True
                        elif elem[i].value < elem[i+2].value:
                            is_the_discrepancies_with_the_3rd_expert = True
                            is_the_discrepancies_with_the_3rd_expert_understatement = True



                elif column_H[i-1].value != None and column_H[i-1].value == column_H[i].value:
                    if elem[i].value == 'X' and elem[i-1].value == 'X':
                        pass
                    elif elem[i].value == 'X' and elem[i-1].value == 0:
                        pass
                    elif elem[i].value == 0 and elem[i-1].value == 'X':
                        pass
                    elif elem[i].value == 'X' and elem[i-1].value != 'X' and elem[i-1].value != 0:
                        is_the_X_mark = True
                        #print(elements, i)
                        #counter += elem[i-1].value #X-задания не должны попадать в другие столбцы
                        
                    elif elem[i].value != 'X' and elem[i].value != 0 and elem[i-1].value == 'X':
                        counter += elem[i].value

                    elif elem[i].value != elem[i-1].value:
                        #print(elem[i].value)
                        #print(elem[i-1].value)
                        counter += abs(elem[i].value - elem[i-1].value)


                    if elem[i+1].value == 'X':
                        pass
                    elif elem[i+1].value != 'X':
                        if elem[i].value == elem[i+1].value:
                            pass
                        elif elem[i].value == 'X':
                            is_the_discrepancies_with_the_3rd_expert = True
                            is_the_unappreciated_tasks = True
                        elif elem[i].value > elem[i+1].value:
                            is_the_discrepancies_with_the_3rd_expert = True
                            is_the_discrepancies_with_the_3rd_expert_overstatement = True
                        elif elem[i].value < elem[i+1].value:
                            is_the_discrepancies_with_the_3rd_expert = True
                            is_the_discrepancies_with_the_3rd_expert_understatement = True




            if is_the_X_mark:
                included_in_the_3rd_check_with_a_discrepancy_X_mark[j] += 1
                is_the_X_mark = False
            else:
                if counter == 2:
                    included_in_the_3rd_check_with_a_discrepancy_2[j] += 1
                elif counter == 3:
                    included_in_the_3rd_check_with_a_discrepancy_3[j] += 1
                elif counter == 4:
                    included_in_the_3rd_check_with_a_discrepancy_4[j] += 1
                elif counter >= 5:
                    included_in_the_3rd_check_with_a_discrepancy_5_and_more[j] += 1
#                elif counter == 1:
#                    print('РАЗНИЦА В 1 БАЛЛ ' + str(i) + elements)

            counter = 0



#            if is_the_discrepancies_with_the_3rd_expert_overstatement and is_the_discrepancies_with_the_3rd_expert_understatement:
#                print('ЕСТЬ И НЕДООЦЕНКА И ПЕРЕОЦЕНКА ' + str(i) + " " + elements)


            if is_the_discrepancies_with_the_3rd_expert:
                discrepancies_with_the_3rd_expert[j] += 1
            is_the_discrepancies_with_the_3rd_expert = False

            if is_the_discrepancies_with_the_3rd_expert_overstatement:
                discrepancies_with_the_3rd_expert_overstatement[j] += 1
            is_the_discrepancies_with_the_3rd_expert_overstatement = False

            if is_the_discrepancies_with_the_3rd_expert_understatement:
                discrepancies_with_the_3rd_expert_understatement[j] += 1
            is_the_discrepancies_with_the_3rd_expert_understatement = False

            if is_the_unappreciated_tasks:
                unappreciated_tasks[j] += 1
            is_the_unappreciated_tasks = False


        #Работы, попавшие на 3-ю проверку (3 эксперт)
        elif elements == column_J[i].value and column_H[i].value != None and column_H[i-2].value != None:
            #included_in_the_3rd_check[j] +=1
            pass



#print(total_amount_without_3)


# print(total_amount)
# print(not_included_in_the_3rd_check)
# print(included_in_the_3rd_check)
# print()

# print(included_in_the_3rd_check_with_a_discrepancy_2)
# print(included_in_the_3rd_check_with_a_discrepancy_3)
# print(included_in_the_3rd_check_with_a_discrepancy_4)
# print(included_in_the_3rd_check_with_a_discrepancy_5_and_more)
# print(included_in_the_3rd_check_with_a_discrepancy_X_mark)
# print()

# print(discrepancies_with_the_3rd_expert)
# print(discrepancies_with_the_3rd_expert_overstatement)
# print(discrepancies_with_the_3rd_expert_understatement)
# print(unappreciated_tasks)
# print()


# print(not_included_in_the_3rd_check)
# print(not_included_in_the_3rd_check_with_discrepancies_1)
# print(not_included_in_the_3rd_check_with_discrepancies_2)
# print(not_included_in_the_3rd_check_with_discrepancies_3_and_more)
# print(not_included_in_the_3rd_check_with_discrepancies_overstatement)
# print(not_included_in_the_3rd_check_with_discrepancies_understatement)



output = openpyxl.Workbook() 

#output = openpyxl.load_workbook("output.xlsx")
Sheet_name = output.sheetnames
list_1 = output[Sheet_name[0]]

list_1.row_dimensions[3].height = 42

list_1.merge_cells('A1:A3')
list_1['A1'] = "№"
list_1.column_dimensions['A'].width = 4
list_1['A1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('B1:B3')
list_1['B1'] = "ФИО"
list_1.column_dimensions['B'].width = 36
list_1['B1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('C1:C3')
list_1['C1'] = "Всего** проверенных работ"
list_1.column_dimensions['C'].width = 10
list_1['C1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('D1:M1')
list_1['D1'] = "Работы, попавшие на 3-ю проверку"
list_1['D1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('D2:D3')
list_1['D2'] = "Всего* (% от Всего**)"
list_1['D2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('E2:I2')
list_1['E2'] = "С расхождением"
list_1['E2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('J2:M2')
list_1['J2'] = "Число несовпадений в оценивании с 3-м экспертом"
list_1['J2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


list_1['E3'] = "В «2» балла  (% от всего*)"
list_1['E3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['F3'] = "В «3» балла (% от всего*)"
list_1['F3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['G3'] = "В «4» балла (% от всего*)"
list_1['G3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['H3'] = "В «5» и более баллов (% от всего*)"
list_1['H3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['I3'] = "«Х» -«балл»  (% от всего*)"
list_1['I3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1['J3'] = "Всего# (% от всего*)"
list_1['J3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['K3'] = "Завышений (% от всего#)"
list_1['K3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['L3'] = "Занижений (% от всего#)"
list_1['L3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['M3'] = "Не оцененных заданий (% от всего#)"
list_1['M3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


list_1.merge_cells('N1:S1')
list_1['N1'] = "Работы, НЕ попавшие на 3-ю проверку"
list_1['N1'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('N2:N3')
list_1['N2'] = "Всего§ (% от Всего**)"
list_1['N2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('O2:Q2')
list_1['O2'] = "С расхождением"
list_1['O2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('R2:R3')
list_1['R2'] = "С завышением по сравнению со 2-м экспертом (% от Всего§)"
list_1['R2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

list_1.merge_cells('S2:S3')
list_1['S2'] = "С занижением по сравнению со 2-м экспертом (% от Всего§)"
list_1['S2'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


list_1['O3'] = "В 1 балл (% от Всего§)"
list_1['O3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['P3'] = "В 2 балла(% от Всего§)"
list_1['P3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
list_1['Q3'] = "В 3 и более балла (% от Всего§)"
list_1['Q3'].alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


columns = ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']
for e in columns:
    list_1.column_dimensions[e].width = 15

for j, elements in enumerate(FIO):
    list_1['A' + str(j+4)] = str(j+1)
    list_1['A' + str(j+4)].style = 'Currency'

    list_1['B' + str(j+4)] = elements #ФИО
    list_1['B' + str(j+4)].style = 'Currency'


#    list_1['C' + str(j+4)] = float(total_amount[j]) #Всего проверенных работ
    list_1['C' + str(j+4)] = str(total_amount[j])
#    list_1['C' + str(j+4)].number_format = BUILTIN_FORMATS[0]
#    list_1['C' + str(j+4)].number_format = numbers.FORMAT_NUMBER
    list_1['C' + str(j+4)].style = 'Currency'
    list_1['C' + str(j+4)].alignment = Alignment(horizontal='center')



    list_1['D' + str(j+4)] = str(included_in_the_3rd_check[j]) + " (" + str(round((included_in_the_3rd_check[j] / total_amount[j] * 100), 1)) + " %)" #Работы, попавшие на 3-ю проверку - Всего* (% от Всего**)
    list_1['D' + str(j+4)].style = 'Currency'
    list_1['D' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, попавшие на 3-ю проверку - С расхождением - В «2» балла  (% от всего*)
    try:
        list_1['E' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_2[j]) + " (" + str(round((included_in_the_3rd_check_with_a_discrepancy_2[j] / included_in_the_3rd_check[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['E' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_2[j]) + " (0.0 %)"
    list_1['E' + str(j+4)].style = 'Currency'
    list_1['E' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, попавшие на 3-ю проверку - С расхождением - В «3» балла (% от всего*)
    try:
        list_1['F' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_3[j]) + " (" + str(round((included_in_the_3rd_check_with_a_discrepancy_3[j] / included_in_the_3rd_check[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['F' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_3[j]) + " (0.0 %)"
    list_1['F' + str(j+4)].style = 'Currency'
    list_1['F' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, попавшие на 3-ю проверку - С расхождением - В «4» балла (% от всего*)
    try:
        list_1['G' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_4[j]) + " (" + str(round((included_in_the_3rd_check_with_a_discrepancy_4[j] / included_in_the_3rd_check[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['G' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_4[j]) + " (0.0 %)"
    list_1['G' + str(j+4)].style = 'Currency'
    list_1['G' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, попавшие на 3-ю проверку - С расхождением - В «5» и более баллов (% от всего*)
    try:
        list_1['H' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_5_and_more[j]) + " (" + str(round((included_in_the_3rd_check_with_a_discrepancy_5_and_more[j] / included_in_the_3rd_check[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['H' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_5_and_more[j]) + " (0.0 %)"
    list_1['H' + str(j+4)].style = 'Currency'
    list_1['H' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, попавшие на 3-ю проверку - С расхождением - «Х» -«балл»  (% от всего*)
    try:
        list_1['I' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_X_mark[j]) + " (" + str(round((included_in_the_3rd_check_with_a_discrepancy_X_mark[j] / included_in_the_3rd_check[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['I' + str(j+4)] = str(included_in_the_3rd_check_with_a_discrepancy_X_mark[j]) + " (0.0 %)"
    list_1['I' + str(j+4)].style = 'Currency'
    list_1['I' + str(j+4)].alignment = Alignment(horizontal='center')


    #Работы, попавшие на 3-ю проверку - Число несовпадений в оценивании с 3-м экспертом - Всего# (% от всего*)
    try:
        list_1['J' + str(j+4)] = str(discrepancies_with_the_3rd_expert[j]) + " (" + str(round((discrepancies_with_the_3rd_expert[j] / included_in_the_3rd_check[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['J' + str(j+4)] = str(discrepancies_with_the_3rd_expert[j]) + " (0.0 %)"
    list_1['J' + str(j+4)].style = 'Currency'
    list_1['J' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, попавшие на 3-ю проверку - Число несовпадений в оценивании с 3-м экспертом - Завышений (% от всего#)
    try:
        list_1['K' + str(j+4)] = str(discrepancies_with_the_3rd_expert_overstatement[j]) + " (" + str(round((discrepancies_with_the_3rd_expert_overstatement[j] / discrepancies_with_the_3rd_expert[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['K' + str(j+4)] = str(discrepancies_with_the_3rd_expert_overstatement[j]) + " (0.0 %)"
    list_1['K' + str(j+4)].style = 'Currency'
    list_1['K' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, попавшие на 3-ю проверку - Число несовпадений в оценивании с 3-м экспертом - Занижений (% от всего#)
    try:
        list_1['L' + str(j+4)] = str(discrepancies_with_the_3rd_expert_understatement[j]) + " (" + str(round((discrepancies_with_the_3rd_expert_understatement[j] / discrepancies_with_the_3rd_expert[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['L' + str(j+4)] = str(discrepancies_with_the_3rd_expert_understatement[j]) + " (0.0 %)"
    list_1['L' + str(j+4)].style = 'Currency'
    list_1['L' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, попавшие на 3-ю проверку - Число несовпадений в оценивании с 3-м экспертом - Не оцененных заданий (% от всего#)
    try:
        list_1['M' + str(j+4)] = str(unappreciated_tasks[j]) + " (" + str(round((unappreciated_tasks[j] / discrepancies_with_the_3rd_expert[j] * 100), 1)) + " %)"
    except ZeroDivisionError:
        list_1['M' + str(j+4)] = str(unappreciated_tasks[j]) + " (0.0 %)"
    list_1['M' + str(j+4)].style = 'Currency'
    list_1['M' + str(j+4)].alignment = Alignment(horizontal='center')


    #Работы, НЕ попавшие на 3-ю проверку - Всего§ (% от Всего**)
    list_1['N' + str(j+4)] = str(not_included_in_the_3rd_check[j]) + " (" + str(round((not_included_in_the_3rd_check[j] / total_amount[j] * 100), 1)) + " %)" #Работы, попавшие на 3-ю проверку - Всего* (% от Всего**)
    list_1['N' + str(j+4)].style = 'Currency'
    list_1['N' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, НЕ попавшие на 3-ю проверку - С расхождением - В 1 балл (% от Всего§)
    try:
        list_1['O' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_1[j]) + " (" + str(round((not_included_in_the_3rd_check_with_discrepancies_1[j] / not_included_in_the_3rd_check[j] * 100), 1)) + " %)" #Работы, попавшие на 3-ю проверку - Всего* (% от Всего**)
    except ZeroDivisionError:
        list_1['O' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_1[j]) + " (0.0 %)"
    list_1['O' + str(j+4)].style = 'Currency'
    list_1['O' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, НЕ попавшие на 3-ю проверку - С расхождением - В 2 балла(% от Всего§)
    try:
        list_1['P' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_2[j]) + " (" + str(round((not_included_in_the_3rd_check_with_discrepancies_2[j] / not_included_in_the_3rd_check[j] * 100), 1)) + " %)" #Работы, попавшие на 3-ю проверку - Всего* (% от Всего**)
    except ZeroDivisionError:
        list_1['P' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_2[j]) + " (0.0 %)"
    list_1['P' + str(j+4)].style = 'Currency'
    list_1['P' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, НЕ попавшие на 3-ю проверку - С расхождением - В 3 и более балла (% от Всего§)
    try:
        list_1['Q' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_3_and_more[j]) + " (" + str(round((not_included_in_the_3rd_check_with_discrepancies_3_and_more[j] / not_included_in_the_3rd_check[j] * 100), 1)) + " %)" #Работы, попавшие на 3-ю проверку - Всего* (% от Всего**)
    except ZeroDivisionError:
        list_1['Q' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_3_and_more[j]) + " (0.0 %)"
    list_1['Q' + str(j+4)].style = 'Currency'
    list_1['Q' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, НЕ попавшие на 3-ю проверку - С завышением по сравнению со 2-м экспертом (% от Всего§)
    try:
        list_1['R' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_overstatement[j]) + " (" + str(round((not_included_in_the_3rd_check_with_discrepancies_overstatement[j] / not_included_in_the_3rd_check[j] * 100), 1)) + " %)" #Работы, попавшие на 3-ю проверку - Всего* (% от Всего**)
    except ZeroDivisionError:
        list_1['R' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_overstatement[j]) + " (0.0 %)"
    list_1['R' + str(j+4)].style = 'Currency'
    list_1['R' + str(j+4)].alignment = Alignment(horizontal='center')

    #Работы, НЕ попавшие на 3-ю проверку - С занижением по сравнению со 2-м экспертом (% от Всего§)
    try:
        list_1['S' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_understatement[j]) + " (" + str(round((not_included_in_the_3rd_check_with_discrepancies_understatement[j] / not_included_in_the_3rd_check[j] * 100), 1)) + " %)" #Работы, попавшие на 3-ю проверку - Всего* (% от Всего**)
    except ZeroDivisionError:
        list_1['S' + str(j+4)] = str(not_included_in_the_3rd_check_with_discrepancies_understatement[j]) + " (0.0 %)"
    list_1['S' + str(j+4)].style = 'Currency'
    list_1['S' + str(j+4)].alignment = Alignment(horizontal='center')


output.save("output.xlsx")


#for key, val in BUILTIN_FORMATS.items():
#    print(f'{key}: {val}')


if 'a' in globals() or input_sheet_name != 'результаты оценивания _Х':
    print("-------------------------")
    print("-------- УСПЕШНО --------")
    print("-------------------------")
    a = input()
